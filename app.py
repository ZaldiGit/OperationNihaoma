from __future__ import annotations

import html
import io
import mimetypes
import re
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

st.set_page_config(page_title="Nihaoma Student Operations System", layout="wide")

APP_TITLE = "Nihaoma Student Operations System"
LOGO_PATH = Path(__file__).parent / "logo-nihaoma-rounded.png"
APPROVAL_PATH = Path(__file__).parent / "approval-composite.png"

REQUIRED_DOC_TYPES = [
    "Paspor",
    "KTP / ID",
    "Ijazah",
    "Transkrip",
    "Sertifikat Bahasa",
    "Foto Formal",
    "Bukti Pembayaran",
    "Surat Pernyataan",
]

STUDENT_COLS = [
    "student_id", "tanggal_input", "nama_lengkap", "nama_panggilan", "jenis_kelamin",
    "tanggal_lahir", "kewarganegaraan", "no_whatsapp", "email", "alamat",
    "no_paspor_atau_nik", "program_diminati", "kampus_tujuan", "kota_tujuan",
    "negara_tujuan", "intake", "durasi_program", "estimasi_biaya", "sumber_leads",
    "pic_admin", "status_proses", "tanggal_follow_up_terakhir", "next_action",
    "tanggal_next_action", "prioritas", "catatan_admin", "is_active"
]

PROGRESS_COLS = [
    "log_id", "student_id", "tanggal_update", "updated_by", "status_lama",
    "status_baru", "catatan", "next_action", "tanggal_next_action"
]

DOC_COLS = [
    "doc_id", "student_id", "nama_mahasiswa", "jenis_dokumen", "nama_file",
    "link_file", "storage_path", "tanggal_upload", "uploaded_by",
    "status_verifikasi", "verified_by", "tanggal_verifikasi", "catatan_verifikasi",
    "versi_dokumen"
]

INVOICE_COLS = [
    "invoice_id", "student_id", "nama_mahasiswa", "kode_invoice", "tanggal_invoice",
    "program", "deskripsi_biaya", "mata_uang", "harga_program", "sudah_dibayar",
    "sisa_tagihan", "status_pelunasan", "status_pengiriman", "tanggal_kirim",
    "bukti_pembayaran_link", "catatan_invoice"
]

PAYMENT_COLS = [
    "payment_id", "invoice_id", "student_id", "tanggal_pembayaran",
    "jumlah_pembayaran", "metode_pembayaran", "bukti_pembayaran_link",
    "dicatat_oleh", "catatan"
]

REFERENCE_SHEETS = {
    "status_proses_ref": [["New Lead"], ["Follow Up"], ["Interested"], ["Dokumen Awal Masuk"], ["Siap Daftar"], ["Sudah Daftar"], ["Menunggu Pembayaran"], ["Proses Visa"], ["Siap Berangkat"], ["Aktif"], ["Selesai"], ["Cancel"]],
    "status_verifikasi_ref": [["Belum Dicek"], ["Valid"], ["Revisi"], ["Tidak Berlaku"]],
    "status_pelunasan_ref": [["Belum Lunas"], ["Sebagian"], ["Lunas"]],
    "status_pengiriman_ref": [["Belum Dikirim"], ["Sudah Dikirim"]],
    "prioritas_ref": [["Tinggi"], ["Sedang"], ["Rendah"]],
    "required_doc_types_ref": [[x] for x in REQUIRED_DOC_TYPES],
}

PROFILE_FIXED = {
    "Nama Brand": "Nihaoma Education Center",
    "Alamat ID": "Gedung Wirausaha Lantai 1,\nJalan HR Rasuna Said Kav. C-5,\nKelurahan Karet, Kecamatan Setia Budi,\nJakarta Selatan 12920",
    "Alamat CN": "佛城西路21号楼 1709-C, 江宁区, 南京市, 江苏, China.",
    "Alamat EN": "Fochengxilu 21 building, room 1709-C, Jiangning District,\nNanjing, Jiangsu, China.",
    "Email": "admin@nihaoma-education.com",
    "Telepon / WA": "+62 812-0000-0000",
    "Info Pembayaran": "Bank BCA - 1234567890 a/n Nihaoma Education Center",
    "Catatan Footer": "Terima kasih atas kepercayaan Anda. Invoice ini diterbitkan untuk kebutuhan administrasi program pendidikan ke China.",
}

STATUS_PROSES_OPTIONS = [r[0] for r in REFERENCE_SHEETS["status_proses_ref"]]
STATUS_VERIF_OPTIONS = [r[0] for r in REFERENCE_SHEETS["status_verifikasi_ref"]]
STATUS_PELUNASAN_OPTIONS = [r[0] for r in REFERENCE_SHEETS["status_pelunasan_ref"]]
STATUS_PENGIRIMAN_OPTIONS = [r[0] for r in REFERENCE_SHEETS["status_pengiriman_ref"]]
PRIORITAS_OPTIONS = [r[0] for r in REFERENCE_SHEETS["prioritas_ref"]]

BLUE = colors.HexColor("#1E88E5")
BLUE_DARK = colors.HexColor("#0F4C81")
ORANGE = colors.HexColor("#F59E0B")
ORANGE_DARK = colors.HexColor("#D97706")
TEXT_DARK = colors.HexColor("#1F2937")
MUTED = colors.HexColor("#6B7280")
LINE = colors.HexColor("#D6DCE5")
LIGHT_BLUE = colors.HexColor("#EEF6FF")
LIGHT_ORANGE = colors.HexColor("#FFF6E8")
WHITE = colors.white


def blank_data():
    return {
        "students": pd.DataFrame(columns=STUDENT_COLS),
        "progress": pd.DataFrame(columns=PROGRESS_COLS),
        "documents": pd.DataFrame(columns=DOC_COLS),
        "invoices": pd.DataFrame(columns=INVOICE_COLS),
        "payments": pd.DataFrame(columns=PAYMENT_COLS),
        "docs_store": {},
        "loaded_source": "template kosong",
    }


def init_state():
    if "nos_data" not in st.session_state:
        st.session_state.nos_data = blank_data()
    if "last_refresh" not in st.session_state:
        st.session_state.last_refresh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def to_datetime_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


def to_float_series(series: pd.Series) -> pd.Series:
    def parse_currency(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return 0.0
        text = str(value).strip()
        if text in {"", "-"}:
            return 0.0
        text = text.replace("Rp", "").replace("rp", "").replace(" ", "")
        if "." in text and "," in text:
            text = text.replace(".", "").replace(",", ".")
        else:
            if text.count(",") > 1:
                text = text.replace(",", "")
            if text.count(".") > 1:
                text = text.replace(".", "")
        text = re.sub(r"[^0-9.\-]", "", text)
        try:
            return float(text) if text not in {"", "-", ".", "-."} else 0.0
        except Exception:
            return 0.0
    return series.apply(parse_currency)


def format_rupiah(value: float) -> str:
    return f"Rp {float(value or 0):,.0f}".replace(",", ".")


def format_date_id(value) -> str:
    if value is None or value == "" or pd.isna(value):
        return "-"
    ts = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(ts):
        return "-"
    return ts.strftime("%d/%m/%Y")


def safe_paragraph_text(value) -> str:
    text = "" if value is None else str(value)
    return html.escape(text).replace("\n", "<br/>")


def next_id(df: pd.DataFrame, prefix: str, width: int = 4, column: str | None = None) -> str:
    if df.empty:
        return f"{prefix}-{1:0{width}d}"
    col = column or df.columns[0]
    nums = df[col].astype(str).str.extract(r"(\d+)$")[0].dropna().astype(int)
    next_num = 1 if nums.empty else int(nums.max()) + 1
    return f"{prefix}-{next_num:0{width}d}"


def recalc_invoice_status(row: pd.Series) -> pd.Series:
    harga = float(row.get("harga_program", 0) or 0)
    paid = float(row.get("sudah_dibayar", 0) or 0)
    balance = max(harga - paid, 0)
    row["sisa_tagihan"] = balance
    if balance <= 0 and harga > 0:
        row["status_pelunasan"] = "Lunas"
    elif paid > 0:
        row["status_pelunasan"] = "Sebagian"
    else:
        row["status_pelunasan"] = "Belum Lunas"
    return row


def normalize_loaded_frames(frames: dict) -> dict:
    data = blank_data()
    for key, cols in [
        ("students", STUDENT_COLS),
        ("progress", PROGRESS_COLS),
        ("documents", DOC_COLS),
        ("invoices", INVOICE_COLS),
        ("payments", PAYMENT_COLS),
    ]:
        df = frames.get(key, pd.DataFrame(columns=cols)).copy()
        for col in cols:
            if col not in df.columns:
                df[col] = None
        df = df[cols]
        data[key] = df

    date_cols = {
        "students": ["tanggal_input", "tanggal_lahir", "tanggal_follow_up_terakhir", "tanggal_next_action"],
        "progress": ["tanggal_update", "tanggal_next_action"],
        "documents": ["tanggal_upload", "tanggal_verifikasi"],
        "invoices": ["tanggal_invoice", "tanggal_kirim"],
        "payments": ["tanggal_pembayaran"],
    }
    float_cols = {
        "students": ["estimasi_biaya"],
        "invoices": ["harga_program", "sudah_dibayar", "sisa_tagihan"],
        "payments": ["jumlah_pembayaran"],
    }
    for key, cols in date_cols.items():
        for col in cols:
            data[key][col] = to_datetime_series(data[key][col])
    for key, cols in float_cols.items():
        for col in cols:
            data[key][col] = to_float_series(data[key][col])

    data["students"] = data["students"].fillna("")
    data["progress"] = data["progress"].fillna("")
    data["documents"] = data["documents"].fillna("")
    data["payments"] = data["payments"].fillna("")
    data["invoices"] = data["invoices"].fillna("")
    if not data["invoices"].empty:
        data["invoices"] = data["invoices"].apply(recalc_invoice_status, axis=1)
    data["docs_store"] = frames.get("docs_store", {})
    data["loaded_source"] = frames.get("loaded_source", "import")
    return data


def workbook_bytes_from_data(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "students_master"
    ws.append(STUDENT_COLS)
    for row in data["students"][STUDENT_COLS].itertuples(index=False):
        ws.append(list(row))

    sheet_map = [
        ("student_progress_log", "progress", PROGRESS_COLS),
        ("student_documents", "documents", DOC_COLS),
        ("student_invoices", "invoices", INVOICE_COLS),
        ("invoice_payment_log", "payments", PAYMENT_COLS),
    ]
    for sheet_name, key, cols in sheet_map:
        ws = wb.create_sheet(sheet_name)
        ws.append(cols)
        export_df = data[key].copy()
        for row in export_df[cols].itertuples(index=False):
            ws.append(list(row))

    ws = wb.create_sheet("README")
    ws["A1"] = "Nihaoma Student Operations System"
    ws["A2"] = "Gunakan workbook ini untuk backup / restore data aplikasi."
    ws["A4"] = "Sheet utama:"
    ws["A5"] = "students_master"
    ws["A6"] = "student_progress_log"
    ws["A7"] = "student_documents"
    ws["A8"] = "student_invoices"
    ws["A9"] = "invoice_payment_log"

    for sheet_name, rows in REFERENCE_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = sheet_name.replace("_ref", "").replace("_", " ").title()
        ws["A2"] = "Value"
        for idx, row in enumerate(rows, start=3):
            ws[f"A{idx}"] = row[0]

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def load_workbook_bytes(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    frames = {}
    mapping = {
        "students_master": ("students", STUDENT_COLS),
        "student_progress_log": ("progress", PROGRESS_COLS),
        "student_documents": ("documents", DOC_COLS),
        "student_invoices": ("invoices", INVOICE_COLS),
        "invoice_payment_log": ("payments", PAYMENT_COLS),
    }
    for sheet_name, (key, cols) in mapping.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.values)
            if rows:
                header = [x if x is not None else "" for x in rows[0]]
                body = rows[1:]
                df = pd.DataFrame(body, columns=header)
            else:
                df = pd.DataFrame(columns=cols)
            frames[key] = df
    return normalize_loaded_frames(frames)


def load_package_bytes(file_bytes: bytes) -> dict:
    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
        names = zf.namelist()
        workbook_name = None
        for candidate in names:
            low = candidate.lower()
            if low.endswith(".xlsx") and ("nihaoma" in low or "data" in low):
                workbook_name = candidate
                break
        if workbook_name is None:
            workbook_name = next((n for n in names if n.lower().endswith(".xlsx")), None)
        if workbook_name is None:
            raise ValueError("Workbook .xlsx tidak ditemukan di package zip.")

        workbook_bytes = zf.read(workbook_name)
        data = load_workbook_bytes(workbook_bytes)
        docs_store = {}
        for name in names:
            if name.endswith("/") or not name.lower().startswith("docs/"):
                continue
            doc_id = Path(name).stem.split("__", 1)[0]
            docs_store[doc_id] = {
                "filename": Path(name).name,
                "bytes": zf.read(name),
                "mime": mimetypes.guess_type(name)[0] or "application/octet-stream",
            }
        data["docs_store"] = docs_store
        data["loaded_source"] = "package zip"
        return data


def export_package_zip(data: dict) -> bytes:
    workbook_bytes = workbook_bytes_from_data(data)
    bio = io.BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("nihaoma_operations_data.xlsx", workbook_bytes)
        for _, row in data["documents"].iterrows():
            doc_id = str(row.get("doc_id", "")).strip()
            student_id = str(row.get("student_id", "")).strip() or "NOID"
            file_info = data["docs_store"].get(doc_id)
            if not file_info:
                continue
            filename = file_info["filename"]
            zf.writestr(f"docs/{student_id}/{doc_id}__{filename}", file_info["bytes"])
    bio.seek(0)
    return bio.read()


def import_controls():
    with st.sidebar.expander("Import / Export Data", expanded=False):
        uploaded = st.file_uploader("Import package (.zip / .xlsx)", type=["zip", "xlsx"], key="import_package")
        col1, col2 = st.columns(2)
        if col1.button("Muat Template Kosong"):
            st.session_state.nos_data = blank_data()
            st.session_state.last_refresh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            st.rerun()
        if uploaded is not None:
            try:
                if uploaded.name.lower().endswith(".zip"):
                    st.session_state.nos_data = load_package_bytes(uploaded.getvalue())
                else:
                    st.session_state.nos_data = load_workbook_bytes(uploaded.getvalue())
                    st.session_state.nos_data["loaded_source"] = "workbook xlsx"
                st.session_state.last_refresh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                st.success("Data berhasil diimport.")
            except Exception as e:
                st.error(f"Gagal import data: {e}")

        st.download_button(
            "Download workbook template (.xlsx)",
            data=workbook_bytes_from_data(blank_data()),
            file_name="nihaoma_operations_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.download_button(
            "Download full package backup (.zip)",
            data=export_package_zip(st.session_state.nos_data),
            file_name="nihaoma_operations_backup.zip",
            mime="application/zip",
        )


def render_header(title: str, subtitle: str = ""):
    col1, col2 = st.columns([1, 6])
    with col1:
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), width=100)
    with col2:
        st.markdown(f"## {title}")
        if subtitle:
            st.write(subtitle)


def dashboard_page(data: dict):
    render_header("Selamat Datang di Dashboard Operasional Nihaoma", "Pantau calon mahasiswa, dokumen, invoice, dan pembayaran dalam satu aplikasi.")
    st.caption(f"Data source: {data.get('loaded_source', '-')} | Last refreshed: {st.session_state.last_refresh}")

    students = data["students"]
    documents = data["documents"]
    invoices = data["invoices"]

    total_students = len(students)
    active_students = int((students.get("is_active", pd.Series(dtype=str)).astype(str).str.upper() == "TRUE").sum()) if not students.empty else 0
    total_invoices = len(invoices)
    outstanding = float(invoices["sisa_tagihan"].sum()) if not invoices.empty else 0

    checklist = []
    if not students.empty:
        for _, stu in students.iterrows():
            sid = str(stu["student_id"])
            stu_docs = documents[documents["student_id"] == sid]["jenis_dokumen"].dropna().astype(str).tolist()
            missing = len([x for x in REQUIRED_DOC_TYPES if x not in stu_docs])
            checklist.append({"student_id": sid, "nama_lengkap": stu["nama_lengkap"], "dokumen_belum_lengkap": missing})
    checklist_df = pd.DataFrame(checklist)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Calon Mahasiswa", total_students)
    c2.metric("Mahasiswa Aktif", active_students)
    c3.metric("Total Invoice", total_invoices)
    c4.metric("Outstanding", format_rupiah(outstanding))

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### Status Proses Mahasiswa")
        if students.empty:
            st.info("Belum ada data mahasiswa.")
        else:
            status_counts = students.groupby("status_proses")["student_id"].count().sort_values(ascending=False)
            st.bar_chart(status_counts)

    with col2:
        st.markdown("### Outstanding per Program")
        if invoices.empty:
            st.info("Belum ada data invoice.")
        else:
            prog = invoices.groupby("program")["sisa_tagihan"].sum().sort_values(ascending=False)
            st.bar_chart(prog)

    col3, col4 = st.columns(2)
    with col3:
        st.markdown("### Dokumen Belum Lengkap")
        if checklist_df.empty:
            st.info("Belum ada data.")
        else:
            st.dataframe(checklist_df.sort_values("dokumen_belum_lengkap", ascending=False).head(10), use_container_width=True, hide_index=True)

    with col4:
        st.markdown("### Invoice Belum Lunas")
        if invoices.empty:
            st.info("Belum ada data.")
        else:
            show = invoices[invoices["status_pelunasan"] != "Lunas"][["kode_invoice", "nama_mahasiswa", "sisa_tagihan", "status_pelunasan"]].copy()
            if show.empty:
                st.success("Semua invoice sudah lunas.")
            else:
                show["sisa_tagihan"] = show["sisa_tagihan"].map(format_rupiah)
                st.dataframe(show.head(10), use_container_width=True, hide_index=True)

    with st.expander("Validasi Data", expanded=False):
        duplicate_ids = students[students["student_id"].duplicated(keep=False)] if not students.empty else pd.DataFrame()
        missing_contacts = students[(students["email"].astype(str) == "") | (students["no_whatsapp"].astype(str) == "")] if not students.empty else pd.DataFrame()
        overpaid = invoices[invoices["sudah_dibayar"] > invoices["harga_program"]] if not invoices.empty else pd.DataFrame()

        a, b, c = st.columns(3)
        a.metric("Student ID duplikat", len(duplicate_ids))
        b.metric("Kontak belum lengkap", len(missing_contacts))
        c.metric("Invoice overpaid", len(overpaid))

        if not duplicate_ids.empty:
            st.warning("Ada student_id duplikat.")
            st.dataframe(duplicate_ids[["student_id", "nama_lengkap"]], use_container_width=True, hide_index=True)
        if not missing_contacts.empty:
            st.warning("Ada mahasiswa dengan email / WhatsApp belum lengkap.")
            st.dataframe(missing_contacts[["student_id", "nama_lengkap", "email", "no_whatsapp"]], use_container_width=True, hide_index=True)
        if not overpaid.empty:
            view = overpaid[["kode_invoice", "nama_mahasiswa", "harga_program", "sudah_dibayar"]].copy()
            view["harga_program"] = view["harga_program"].map(format_rupiah)
            view["sudah_dibayar"] = view["sudah_dibayar"].map(format_rupiah)
            st.warning("Ada invoice dengan pembayaran melebihi nilai tagihan.")
            st.dataframe(view, use_container_width=True, hide_index=True)


def students_page(data: dict):
    render_header("Modul Calon Mahasiswa", "Input data, lihat detail, dan kelola progress mahasiswa.")
    tabs = st.tabs(["Daftar Mahasiswa", "Tambah Data", "Detail & Progress"])

    students = data["students"]
    progress = data["progress"]

    with tabs[0]:
        search = st.text_input("Cari mahasiswa")
        col1, col2, col3 = st.columns(3)
        program_options = sorted([x for x in students["program_diminati"].astype(str).unique().tolist() if x]) if not students.empty else []
        status_options = sorted([x for x in students["status_proses"].astype(str).unique().tolist() if x]) if not students.empty else []
        pic_options = sorted([x for x in students["pic_admin"].astype(str).unique().tolist() if x]) if not students.empty else []
        selected_program = col1.multiselect("Program", program_options, default=program_options)
        selected_status = col2.multiselect("Status Proses", status_options, default=status_options)
        selected_pic = col3.multiselect("PIC", pic_options, default=pic_options)

        show = students.copy()
        if not show.empty:
            if search:
                kw = search.lower()
                show = show[
                    show["nama_lengkap"].astype(str).str.lower().str.contains(kw, na=False)
                    | show["student_id"].astype(str).str.lower().str.contains(kw, na=False)
                    | show["email"].astype(str).str.lower().str.contains(kw, na=False)
                ]
            if selected_program:
                show = show[show["program_diminati"].isin(selected_program)]
            if selected_status:
                show = show[show["status_proses"].isin(selected_status)]
            if selected_pic:
                show = show[show["pic_admin"].isin(selected_pic)]
            display = show[["student_id", "nama_lengkap", "program_diminati", "intake", "pic_admin", "status_proses", "tanggal_input"]].copy()
            display["tanggal_input"] = display["tanggal_input"].apply(format_date_id)
            st.dataframe(display, use_container_width=True, hide_index=True)
        else:
            st.info("Belum ada data mahasiswa.")

    with tabs[1]:
        with st.form("add_student_form"):
            c1, c2, c3 = st.columns(3)
            nama_lengkap = c1.text_input("Nama Lengkap *")
            nama_panggilan = c2.text_input("Nama Panggilan")
            jenis_kelamin = c3.selectbox("Jenis Kelamin", ["", "Pria", "Wanita"])
            c4, c5, c6 = st.columns(3)
            tanggal_lahir = c4.date_input("Tanggal Lahir", value=None)
            kewarganegaraan = c5.text_input("Kewarganegaraan", value="Indonesia")
            no_whatsapp = c6.text_input("No. WhatsApp")
            c7, c8, c9 = st.columns(3)
            email = c7.text_input("Email")
            alamat = c8.text_input("Alamat")
            no_paspor_atau_nik = c9.text_input("No. Paspor / NIK")
            c10, c11, c12 = st.columns(3)
            program_diminati = c10.text_input("Program Diminati")
            kampus_tujuan = c11.text_input("Kampus Tujuan")
            kota_tujuan = c12.text_input("Kota Tujuan")
            c13, c14, c15 = st.columns(3)
            negara_tujuan = c13.text_input("Negara Tujuan", value="China")
            intake = c14.text_input("Intake")
            durasi_program = c15.text_input("Durasi Program")
            c16, c17, c18 = st.columns(3)
            estimasi_biaya = c16.number_input("Estimasi Biaya", min_value=0.0, step=100000.0)
            sumber_leads = c17.text_input("Sumber Leads")
            pic_admin = c18.text_input("PIC Admin")
            c19, c20, c21 = st.columns(3)
            status_proses = c19.selectbox("Status Proses", STATUS_PROSES_OPTIONS, index=0)
            prioritas = c20.selectbox("Prioritas", PRIORITAS_OPTIONS, index=1)
            is_active = c21.checkbox("Aktif", value=True)
            c22, c23 = st.columns(2)
            next_action = c22.text_input("Next Action")
            tanggal_next_action = c23.date_input("Tanggal Next Action", value=None)
            catatan_admin = st.text_area("Catatan Admin")
            submitted = st.form_submit_button("Simpan Mahasiswa")

        if submitted:
            if not nama_lengkap.strip():
                st.error("Nama Lengkap wajib diisi.")
            else:
                new_id = next_id(students, "STD", column="student_id")
                new_row = {
                    "student_id": new_id,
                    "tanggal_input": pd.Timestamp(datetime.now().date()),
                    "nama_lengkap": nama_lengkap,
                    "nama_panggilan": nama_panggilan,
                    "jenis_kelamin": jenis_kelamin,
                    "tanggal_lahir": pd.Timestamp(tanggal_lahir) if tanggal_lahir else pd.NaT,
                    "kewarganegaraan": kewarganegaraan,
                    "no_whatsapp": no_whatsapp,
                    "email": email,
                    "alamat": alamat,
                    "no_paspor_atau_nik": no_paspor_atau_nik,
                    "program_diminati": program_diminati,
                    "kampus_tujuan": kampus_tujuan,
                    "kota_tujuan": kota_tujuan,
                    "negara_tujuan": negara_tujuan,
                    "intake": intake,
                    "durasi_program": durasi_program,
                    "estimasi_biaya": estimasi_biaya,
                    "sumber_leads": sumber_leads,
                    "pic_admin": pic_admin,
                    "status_proses": status_proses,
                    "tanggal_follow_up_terakhir": pd.NaT,
                    "next_action": next_action,
                    "tanggal_next_action": pd.Timestamp(tanggal_next_action) if tanggal_next_action else pd.NaT,
                    "prioritas": prioritas,
                    "catatan_admin": catatan_admin,
                    "is_active": bool(is_active),
                }
                st.session_state.nos_data["students"] = pd.concat([students, pd.DataFrame([new_row])], ignore_index=True)
                log_row = {
                    "log_id": next_id(progress, "LOG", column="log_id"),
                    "student_id": new_id,
                    "tanggal_update": pd.Timestamp(datetime.now()),
                    "updated_by": pic_admin or "Admin",
                    "status_lama": "",
                    "status_baru": status_proses,
                    "catatan": catatan_admin,
                    "next_action": next_action,
                    "tanggal_next_action": pd.Timestamp(tanggal_next_action) if tanggal_next_action else pd.NaT,
                }
                st.session_state.nos_data["progress"] = pd.concat([progress, pd.DataFrame([log_row])], ignore_index=True)
                st.success(f"Mahasiswa {nama_lengkap} berhasil disimpan dengan ID {new_id}.")
                st.rerun()

    with tabs[2]:
        if students.empty:
            st.info("Belum ada data mahasiswa.")
        else:
            student_options = students["student_id"].astype(str) + " - " + students["nama_lengkap"].astype(str)
            selected = st.selectbox("Pilih Mahasiswa", student_options)
            selected_id = selected.split(" - ", 1)[0]
            row = students[students["student_id"] == selected_id].iloc[0]

            st.markdown(f"### {row['nama_lengkap']}")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Student ID", row["student_id"])
            c2.metric("Program", row["program_diminati"] or "-")
            c3.metric("Status", row["status_proses"] or "-")
            c4.metric("PIC", row["pic_admin"] or "-")

            t1, t2, t3 = st.tabs(["Profil", "Update Progress", "Riwayat"])
            with t1:
                show = pd.DataFrame({
                    "Field": [
                        "Nama Lengkap", "No. WhatsApp", "Email", "Paspor / NIK", "Program",
                        "Intake", "Negara Tujuan", "Next Action", "Tanggal Next Action", "Catatan"
                    ],
                    "Value": [
                        row["nama_lengkap"], row["no_whatsapp"], row["email"], row["no_paspor_atau_nik"], row["program_diminati"],
                        row["intake"], row["negara_tujuan"], row["next_action"], format_date_id(row["tanggal_next_action"]), row["catatan_admin"]
                    ]
                })
                st.dataframe(show, use_container_width=True, hide_index=True)

            with t2:
                with st.form("update_progress_form"):
                    current_status = row["status_proses"] if row["status_proses"] in STATUS_PROSES_OPTIONS else STATUS_PROSES_OPTIONS[0]
                    new_status = st.selectbox("Status Baru", STATUS_PROSES_OPTIONS, index=STATUS_PROSES_OPTIONS.index(current_status))
                    new_next_action = st.text_input("Next Action", value=str(row["next_action"]))
                    current_next_date = row["tanggal_next_action"].date() if pd.notna(row["tanggal_next_action"]) else None
                    new_next_date = st.date_input("Tanggal Next Action", value=current_next_date)
                    new_note = st.text_area("Catatan Update")
                    updated_by = st.text_input("Updated By", value=str(row["pic_admin"] or "Admin"))
                    submitted_update = st.form_submit_button("Simpan Update")
                if submitted_update:
                    idx = students.index[students["student_id"] == selected_id][0]
                    st.session_state.nos_data["students"].loc[idx, "status_proses"] = new_status
                    st.session_state.nos_data["students"].loc[idx, "next_action"] = new_next_action
                    st.session_state.nos_data["students"].loc[idx, "tanggal_next_action"] = pd.Timestamp(new_next_date) if new_next_date else pd.NaT
                    st.session_state.nos_data["students"].loc[idx, "tanggal_follow_up_terakhir"] = pd.Timestamp(datetime.now())
                    if new_note:
                        existing_note = str(st.session_state.nos_data["students"].loc[idx, "catatan_admin"] or "")
                        combined = (existing_note + "\n" + new_note).strip() if existing_note else new_note
                        st.session_state.nos_data["students"].loc[idx, "catatan_admin"] = combined

                    new_log = {
                        "log_id": next_id(progress, "LOG", column="log_id"),
                        "student_id": selected_id,
                        "tanggal_update": pd.Timestamp(datetime.now()),
                        "updated_by": updated_by or "Admin",
                        "status_lama": row["status_proses"],
                        "status_baru": new_status,
                        "catatan": new_note,
                        "next_action": new_next_action,
                        "tanggal_next_action": pd.Timestamp(new_next_date) if new_next_date else pd.NaT,
                    }
                    st.session_state.nos_data["progress"] = pd.concat([progress, pd.DataFrame([new_log])], ignore_index=True)
                    st.success("Progress mahasiswa berhasil diperbarui.")
                    st.rerun()

            with t3:
                logs = progress[progress["student_id"] == selected_id].copy()
                if logs.empty:
                    st.info("Belum ada riwayat.")
                else:
                    logs["tanggal_update"] = logs["tanggal_update"].apply(format_date_id)
                    st.dataframe(logs[["tanggal_update", "updated_by", "status_lama", "status_baru", "catatan", "next_action"]], use_container_width=True, hide_index=True)


def documents_page(data: dict):
    render_header("Modul Dokumen", "Upload dokumen per mahasiswa, cek checklist, dan download file.")
    tabs = st.tabs(["Upload Dokumen", "Checklist", "Daftar & Download"])

    students = data["students"]
    documents = data["documents"]

    with tabs[0]:
        if students.empty:
            st.info("Tambahkan data mahasiswa terlebih dahulu.")
        else:
            student_options = students["student_id"].astype(str) + " - " + students["nama_lengkap"].astype(str)
            with st.form("upload_doc_form"):
                selected = st.selectbox("Pilih Mahasiswa", student_options)
                selected_id = selected.split(" - ", 1)[0]
                selected_name = selected.split(" - ", 1)[1]
                jenis_dokumen = st.selectbox("Jenis Dokumen", REQUIRED_DOC_TYPES + ["Form Aplikasi", "Dokumen Visa", "Lainnya"])
                versi = st.text_input("Versi Dokumen", value="v1")
                uploaded_by = st.text_input("Uploaded By", value="Admin")
                status_verifikasi = st.selectbox("Status Verifikasi", STATUS_VERIF_OPTIONS, index=0)
                catatan = st.text_area("Catatan Verifikasi / Upload")
                file = st.file_uploader("Upload File", type=None)
                submitted = st.form_submit_button("Upload Dokumen")

            if submitted:
                if file is None:
                    st.error("File wajib dipilih.")
                else:
                    doc_id = next_id(documents, "DOC", column="doc_id")
                    storage_path = f"docs/{selected_id}/{doc_id}__{file.name}"
                    st.session_state.nos_data["docs_store"][doc_id] = {
                        "filename": file.name,
                        "bytes": file.getvalue(),
                        "mime": file.type or mimetypes.guess_type(file.name)[0] or "application/octet-stream",
                    }
                    row = {
                        "doc_id": doc_id,
                        "student_id": selected_id,
                        "nama_mahasiswa": selected_name,
                        "jenis_dokumen": jenis_dokumen,
                        "nama_file": file.name,
                        "link_file": "",
                        "storage_path": storage_path,
                        "tanggal_upload": pd.Timestamp(datetime.now()),
                        "uploaded_by": uploaded_by,
                        "status_verifikasi": status_verifikasi,
                        "verified_by": "",
                        "tanggal_verifikasi": pd.NaT,
                        "catatan_verifikasi": catatan,
                        "versi_dokumen": versi,
                    }
                    st.session_state.nos_data["documents"] = pd.concat([documents, pd.DataFrame([row])], ignore_index=True)
                    st.success(f"Dokumen {file.name} berhasil diupload.")
                    st.rerun()

    with tabs[1]:
        if students.empty:
            st.info("Belum ada data mahasiswa.")
        else:
            rows = []
            for _, stu in students.iterrows():
                sid = str(stu["student_id"])
                stu_docs = documents[documents["student_id"] == sid]["jenis_dokumen"].astype(str).tolist()
                row = {"student_id": sid, "nama_mahasiswa": stu["nama_lengkap"]}
                for dtype in REQUIRED_DOC_TYPES:
                    row[dtype] = "Ada" if dtype in stu_docs else "Belum"
                row["Total Missing"] = sum(1 for d in REQUIRED_DOC_TYPES if d not in stu_docs)
                rows.append(row)
            pivot = pd.DataFrame(rows).sort_values("Total Missing", ascending=False)
            st.dataframe(pivot, use_container_width=True, hide_index=True)

    with tabs[2]:
        if documents.empty:
            st.info("Belum ada dokumen.")
        else:
            show = documents.copy()
            show["tanggal_upload"] = show["tanggal_upload"].apply(format_date_id)
            st.dataframe(show[["doc_id", "student_id", "nama_mahasiswa", "jenis_dokumen", "nama_file", "tanggal_upload", "status_verifikasi"]], use_container_width=True, hide_index=True)

            st.markdown("### Download Dokumen")
            doc_options = show["doc_id"].astype(str) + " - " + show["nama_file"].astype(str)
            selected_doc = st.selectbox("Pilih Dokumen", doc_options)
            doc_id = selected_doc.split(" - ", 1)[0]
            file_info = data["docs_store"].get(doc_id)
            row = show[show["doc_id"] == doc_id].iloc[0]
            if file_info:
                st.download_button("Download Dokumen", data=file_info["bytes"], file_name=file_info["filename"], mime=file_info["mime"])
                st.caption(f"Milik: {row['nama_mahasiswa']} | Jenis: {row['jenis_dokumen']}")
            else:
                st.warning("Binary file belum tersedia. Gunakan package backup untuk menyimpan dan restore file dokumen.")


def register_fonts():
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    except Exception:
        pass


def generate_invoice_pdf(invoice_row: dict) -> bytes:
    register_fonts()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm)

    styles = getSampleStyleSheet()
    brand_name = ParagraphStyle("BrandName", parent=styles["Title"], fontSize=20, leading=23, textColor=TEXT_DARK)
    invoice_title = ParagraphStyle("InvoiceTitle", parent=styles["Title"], fontSize=23, leading=26, textColor=ORANGE_DARK, alignment=1)
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=9.2, leading=12, textColor=TEXT_DARK)
    chinese = ParagraphStyle("Chinese", parent=body, fontName="STSong-Light")
    label = ParagraphStyle("Label", parent=body, fontSize=9, leading=12)
    value = ParagraphStyle("Value", parent=body, fontSize=9.4, leading=12)

    story = []
    accent = Table([["", "", ""]], colWidths=[112 * mm, 40 * mm, 28 * mm], rowHeights=[4 * mm])
    accent.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), BLUE_DARK),
        ("BACKGROUND", (1, 0), (1, 0), BLUE),
        ("BACKGROUND", (2, 0), (2, 0), ORANGE),
    ]))
    story.append(accent)
    story.append(Spacer(1, 6))

    logo_flowable = RLImage(str(LOGO_PATH), width=28 * mm, height=28 * mm) if LOGO_PATH.exists() else Paragraph("", body)
    identity_table = Table([
        [logo_flowable, Paragraph(f"<b>{safe_paragraph_text(PROFILE_FIXED['Nama Brand'])}</b>", brand_name)],
        ["", Paragraph(safe_paragraph_text(PROFILE_FIXED["Alamat ID"]), body)],
        ["", Paragraph(safe_paragraph_text(PROFILE_FIXED["Alamat CN"]), chinese)],
        ["", Paragraph(safe_paragraph_text(PROFILE_FIXED["Alamat EN"]), body)],
        ["", Paragraph(f"Email: {safe_paragraph_text(PROFILE_FIXED['Email'])}<br/>WhatsApp: {safe_paragraph_text(PROFILE_FIXED['Telepon / WA'])}", body)],
    ], colWidths=[32 * mm, 78 * mm])
    identity_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("VALIGN", (0, 0), (0, 0), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("SPAN", (0, 1), (0, 4)),
    ]))

    invoice_card = Table([
        [Paragraph("INVOICE", invoice_title)],
        [Paragraph(f"<b>Kode</b>: {safe_paragraph_text(invoice_row.get('kode_invoice', '-'))}", body)],
        [Paragraph(f"<b>Tanggal</b>: {format_date_id(invoice_row.get('tanggal_invoice'))}", body)],
        [Paragraph(f"<b>Status</b>: {safe_paragraph_text(invoice_row.get('status_pelunasan', '-'))}", body)],
    ], colWidths=[64 * mm])
    invoice_card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT_ORANGE),
        ("BOX", (0, 0), (-1, -1), 0.9, ORANGE),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))

    top = Table([[identity_table, invoice_card]], colWidths=[112 * mm, 66 * mm])
    top.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    story.append(top)
    story.append(Spacer(1, 8))

    def section_title(text, color):
        return Table([[Paragraph(f"<font color='white'><b>{safe_paragraph_text(text)}</b></font>", styles["BodyText"])]], colWidths=[178 * mm], rowHeights=[8 * mm], style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), color), ("LEFTPADDING", (0, 0), (-1, -1), 8)]))

    story.append(section_title("Student & Invoice Details", BLUE))
    detail_rows = [
        [Paragraph("<b>Nama Student</b>", label), Paragraph(safe_paragraph_text(invoice_row.get("nama_mahasiswa", "-")), value), Paragraph("<b>Program</b>", label), Paragraph(safe_paragraph_text(invoice_row.get("program", "-")), value)],
        [Paragraph("<b>Kode Invoice</b>", label), Paragraph(safe_paragraph_text(invoice_row.get("kode_invoice", "-")), value), Paragraph("<b>Status Pelunasan</b>", label), Paragraph(safe_paragraph_text(invoice_row.get("status_pelunasan", "-")), value)],
        [Paragraph("<b>Tanggal Invoice</b>", label), Paragraph(format_date_id(invoice_row.get("tanggal_invoice")), value), Paragraph("<b>Status Pengiriman</b>", label), Paragraph(safe_paragraph_text(invoice_row.get("status_pengiriman", "-")), value)],
        [Paragraph("<b>Deskripsi</b>", label), Paragraph(safe_paragraph_text(invoice_row.get("deskripsi_biaya", "-")), value), Paragraph("<b>Mata Uang</b>", label), Paragraph(safe_paragraph_text(invoice_row.get("mata_uang", "IDR")), value)],
    ]
    detail_table = Table(detail_rows, colWidths=[30 * mm, 58 * mm, 33 * mm, 57 * mm])
    detail_table.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [WHITE, LIGHT_BLUE]),
        ("GRID", (0, 0), (-1, -1), 0.45, LINE),
        ("BOX", (0, 0), (-1, -1), 0.9, BLUE),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(detail_table)
    story.append(Spacer(1, 8))

    story.append(section_title("Program Charge", ORANGE))
    charge_rows = [
        [Paragraph("<font color='white'><b>No</b></font>", label), Paragraph("<font color='white'><b>Deskripsi</b></font>", label), Paragraph("<font color='white'><b>Mata Uang</b></font>", label), Paragraph("<font color='white'><b>Total</b></font>", label)],
        [Paragraph("1", value), Paragraph(safe_paragraph_text(invoice_row.get("deskripsi_biaya", "-")), value), Paragraph(safe_paragraph_text(invoice_row.get("mata_uang", "IDR")), value), Paragraph(format_rupiah(invoice_row.get("harga_program", 0)), value)],
    ]
    charge_table = Table(charge_rows, colWidths=[14 * mm, 110 * mm, 18 * mm, 36 * mm])
    charge_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ORANGE),
        ("BACKGROUND", (0, 1), (-1, -1), LIGHT_ORANGE),
        ("GRID", (0, 0), (-1, -1), 0.45, LINE),
        ("BOX", (0, 0), (-1, -1), 0.9, ORANGE_DARK),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(charge_table)
    story.append(Spacer(1, 8))

    story.append(section_title("Pembayaran & Ringkasan", BLUE_DARK))
    summary_text = (
        f"<b>Total Program</b>: {format_rupiah(invoice_row.get('harga_program', 0))}<br/>"
        f"<b>Sudah Dibayar</b>: {format_rupiah(invoice_row.get('sudah_dibayar', 0))}<br/>"
        f"<b>Sisa Tagihan</b>: {format_rupiah(invoice_row.get('sisa_tagihan', 0))}"
    )
    payment_table = Table([[Paragraph(safe_paragraph_text(PROFILE_FIXED.get("Info Pembayaran", "-")), value), Paragraph(summary_text, value)]], colWidths=[112 * mm, 66 * mm])
    payment_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), LIGHT_BLUE),
        ("BACKGROUND", (1, 0), (1, 0), LIGHT_ORANGE),
        ("BOX", (0, 0), (-1, -1), 0.9, BLUE_DARK),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, LINE),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(payment_table)
    story.append(Spacer(1, 8))

    footer_note = Table([[Paragraph(safe_paragraph_text(PROFILE_FIXED.get("Catatan Footer", "")), body)]], colWidths=[118 * mm])
    footer_note.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7E6")),
        ("BOX", (0, 0), (-1, -1), 0.9, ORANGE),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    signature_title = ParagraphStyle("SignatureTitle", parent=body, fontSize=8.8, leading=11, textColor=MUTED)
    signature_name = ParagraphStyle("SignatureName", parent=body, fontSize=10, leading=12, textColor=TEXT_DARK, alignment=1)
    signature_role = ParagraphStyle("SignatureRole", parent=body, fontSize=9, leading=11, textColor=MUTED, alignment=1)

    if APPROVAL_PATH.exists():
        approval_img = RLImage(str(APPROVAL_PATH), width=56 * mm, height=20 * mm)
    else:
        approval_img = Paragraph("", body)

    approval_block = Table(
        [
            [Paragraph("<b>Authorized Signature</b>", signature_title)],
            [approval_img],
            [Paragraph("<b>Yenny Pricila</b>", signature_name)],
            [Paragraph("Management", signature_role)],
        ],
        colWidths=[60 * mm],
    )
    approval_block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("LINEABOVE", (0, 0), (-1, 0), 0.4, LINE),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("ALIGN", (0, 1), (0, 1), "CENTER"),
    ]))

    footer_combo = Table([[footer_note, approval_block]], colWidths=[118 * mm, 60 * mm])
    footer_combo.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "BOTTOM"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    story.append(footer_combo)

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def invoices_page(data: dict):
    render_header("Modul Invoice & Pembayaran", "Kelola invoice, catat pembayaran, dan unduh PDF invoice.")
    tabs = st.tabs(["Dashboard Invoice", "Tambah Invoice", "Catat Pembayaran", "PDF Invoice"])

    students = data["students"]
    invoices = data["invoices"]
    payments = data["payments"]

    with tabs[0]:
        c1, c2, c3, c4 = st.columns(4)
        total_invoice = len(invoices)
        total_nilai = float(invoices["harga_program"].sum()) if not invoices.empty else 0
        total_paid = float(invoices["sudah_dibayar"].sum()) if not invoices.empty else 0
        total_outstanding = float(invoices["sisa_tagihan"].sum()) if not invoices.empty else 0
        c1.metric("Total Invoice", total_invoice)
        c2.metric("Total Nilai", format_rupiah(total_nilai))
        c3.metric("Sudah Dibayar", format_rupiah(total_paid))
        c4.metric("Outstanding", format_rupiah(total_outstanding))

        if not invoices.empty:
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("### Status Pelunasan")
                st.bar_chart(invoices.groupby("status_pelunasan")["invoice_id"].count().sort_values(ascending=False))
            with col2:
                st.markdown("### Outstanding per Program")
                st.bar_chart(invoices.groupby("program")["sisa_tagihan"].sum().sort_values(ascending=False))

            show = invoices.copy()
            show["tanggal_invoice"] = show["tanggal_invoice"].apply(format_date_id)
            show["harga_program"] = show["harga_program"].map(format_rupiah)
            show["sudah_dibayar"] = show["sudah_dibayar"].map(format_rupiah)
            show["sisa_tagihan"] = show["sisa_tagihan"].map(format_rupiah)
            st.markdown("### Daftar Invoice")
            st.dataframe(show[["kode_invoice", "nama_mahasiswa", "program", "tanggal_invoice", "harga_program", "sudah_dibayar", "sisa_tagihan", "status_pelunasan", "status_pengiriman"]], use_container_width=True, hide_index=True)
        else:
            st.info("Belum ada invoice.")

    with tabs[1]:
        if students.empty:
            st.info("Tambahkan data mahasiswa terlebih dahulu.")
        else:
            student_options = students["student_id"].astype(str) + " - " + students["nama_lengkap"].astype(str)
            with st.form("add_invoice_form"):
                selected = st.selectbox("Pilih Mahasiswa", student_options)
                selected_id = selected.split(" - ", 1)[0]
                selected_name = selected.split(" - ", 1)[1]
                selected_student = students[students["student_id"] == selected_id].iloc[0]
                c1, c2 = st.columns(2)
                program = c1.text_input("Program", value=str(selected_student["program_diminati"]))
                deskripsi = c2.text_input("Deskripsi Biaya", value=f"Biaya program {selected_student['program_diminati']}")
                c3, c4, c5 = st.columns(3)
                mata_uang = c3.text_input("Mata Uang", value="IDR")
                harga_program = c4.number_input("Harga Program", min_value=0.0, value=float(selected_student["estimasi_biaya"] or 0), step=100000.0)
                sudah_dibayar = c5.number_input("Sudah Dibayar", min_value=0.0, value=0.0, step=100000.0)
                c6, c7, c8 = st.columns(3)
                tanggal_invoice = c6.date_input("Tanggal Invoice", value=datetime.now().date())
                status_pengiriman = c7.selectbox("Status Pengiriman", STATUS_PENGIRIMAN_OPTIONS, index=0)
                tanggal_kirim = c8.date_input("Tanggal Kirim", value=None)
                catatan_invoice = st.text_area("Catatan Invoice")
                submitted = st.form_submit_button("Simpan Invoice")
            if submitted:
                invoice_id = next_id(invoices, "INVROW", column="invoice_id")
                kode_invoice = next_id(invoices, "NHEC", column="kode_invoice")
                row = pd.Series({
                    "invoice_id": invoice_id,
                    "student_id": selected_id,
                    "nama_mahasiswa": selected_name,
                    "kode_invoice": kode_invoice,
                    "tanggal_invoice": pd.Timestamp(tanggal_invoice),
                    "program": program,
                    "deskripsi_biaya": deskripsi,
                    "mata_uang": mata_uang,
                    "harga_program": float(harga_program),
                    "sudah_dibayar": float(sudah_dibayar),
                    "sisa_tagihan": 0.0,
                    "status_pelunasan": "Belum Lunas",
                    "status_pengiriman": status_pengiriman,
                    "tanggal_kirim": pd.Timestamp(tanggal_kirim) if tanggal_kirim else pd.NaT,
                    "bukti_pembayaran_link": "",
                    "catatan_invoice": catatan_invoice,
                })
                row = recalc_invoice_status(row)
                st.session_state.nos_data["invoices"] = pd.concat([invoices, pd.DataFrame([row])], ignore_index=True)
                st.success(f"Invoice {kode_invoice} berhasil dibuat.")
                st.rerun()

    with tabs[2]:
        if invoices.empty:
            st.info("Belum ada invoice.")
        else:
            invoice_options = invoices["kode_invoice"].astype(str) + " - " + invoices["nama_mahasiswa"].astype(str)
            with st.form("payment_form"):
                selected = st.selectbox("Pilih Invoice", invoice_options)
                invoice_code = selected.split(" - ", 1)[0]
                row = invoices[invoices["kode_invoice"] == invoice_code].iloc[0]
                st.caption(f"Outstanding saat ini: {format_rupiah(row['sisa_tagihan'])}")
                c1, c2, c3 = st.columns(3)
                tanggal_pembayaran = c1.date_input("Tanggal Pembayaran", value=datetime.now().date())
                jumlah = c2.number_input("Jumlah Pembayaran", min_value=0.0, step=100000.0)
                metode = c3.text_input("Metode Pembayaran", value="Transfer")
                c4, c5 = st.columns(2)
                dicatat_oleh = c4.text_input("Dicatat Oleh", value="Finance")
                bukti_link = c5.text_input("Link Bukti Pembayaran")
                catatan = st.text_area("Catatan Pembayaran")
                submitted_payment = st.form_submit_button("Simpan Pembayaran")

            if submitted_payment:
                payment_id = next_id(payments, "PAY", column="payment_id")
                new_payment = {
                    "payment_id": payment_id,
                    "invoice_id": row["invoice_id"],
                    "student_id": row["student_id"],
                    "tanggal_pembayaran": pd.Timestamp(tanggal_pembayaran),
                    "jumlah_pembayaran": float(jumlah),
                    "metode_pembayaran": metode,
                    "bukti_pembayaran_link": bukti_link,
                    "dicatat_oleh": dicatat_oleh,
                    "catatan": catatan,
                }
                st.session_state.nos_data["payments"] = pd.concat([payments, pd.DataFrame([new_payment])], ignore_index=True)

                idx = invoices.index[invoices["kode_invoice"] == invoice_code][0]
                st.session_state.nos_data["invoices"].loc[idx, "sudah_dibayar"] = float(invoices.loc[idx, "sudah_dibayar"] or 0) + float(jumlah)
                st.session_state.nos_data["invoices"].loc[idx, "bukti_pembayaran_link"] = bukti_link
                updated_row = recalc_invoice_status(st.session_state.nos_data["invoices"].loc[idx].copy())
                for col in ["sisa_tagihan", "status_pelunasan", "sudah_dibayar"]:
                    st.session_state.nos_data["invoices"].loc[idx, col] = updated_row[col]
                st.success("Pembayaran berhasil dicatat.")
                st.rerun()

            st.markdown("### Histori Pembayaran")
            show = payments.copy()
            if show.empty:
                st.info("Belum ada histori pembayaran.")
            else:
                show["tanggal_pembayaran"] = show["tanggal_pembayaran"].apply(format_date_id)
                show["jumlah_pembayaran"] = show["jumlah_pembayaran"].map(format_rupiah)
                st.dataframe(show[["payment_id", "invoice_id", "student_id", "tanggal_pembayaran", "jumlah_pembayaran", "metode_pembayaran", "dicatat_oleh"]], use_container_width=True, hide_index=True)

    with tabs[3]:
        if invoices.empty:
            st.info("Belum ada invoice.")
        else:
            invoice_options = invoices["kode_invoice"].astype(str) + " - " + invoices["nama_mahasiswa"].astype(str)
            selected = st.selectbox("Pilih Invoice untuk PDF", invoice_options)
            invoice_code = selected.split(" - ", 1)[0]
            row = invoices[invoices["kode_invoice"] == invoice_code].iloc[0].to_dict()
            c1, c2, c3 = st.columns(3)
            c1.metric("Nama Mahasiswa", row.get("nama_mahasiswa", "-"))
            c2.metric("Program", row.get("program", "-"))
            c3.metric("Sisa Tagihan", format_rupiah(row.get("sisa_tagihan", 0)))
            pdf_bytes = generate_invoice_pdf(row)
            st.download_button("Download PDF Invoice", data=pdf_bytes, file_name=f"{invoice_code}.pdf", mime="application/pdf")


def help_page():
    render_header("Bantuan & SOP", "Panduan singkat untuk tim operasional Nihaoma.")
    st.markdown("### Cara Memulai")
    st.markdown(
        "1. Gunakan **Download workbook template** atau **Muat Template Kosong** dari sidebar.\n"
        "2. Tambahkan data di modul **Calon Mahasiswa**.\n"
        "3. Upload dokumen di modul **Dokumen**.\n"
        "4. Buat invoice di modul **Invoice & Pembayaran**.\n"
        "5. Backup data secara berkala dengan **Download full package backup (.zip)**."
    )
    st.markdown("### Catatan Penting")
    st.info(
        "Versi ini dirancang agar bisa jalan di Streamlit tanpa backend tambahan. "
        "Dokumen yang diupload disimpan di session aplikasi dan bisa dipersist dengan export package backup (.zip)."
    )


def sidebar_navigation():
    import_controls()
    return st.sidebar.radio("Menu", ["Dashboard", "Calon Mahasiswa", "Dokumen", "Invoice & Pembayaran", "Bantuan & SOP"])


def main():
    init_state()
    data = st.session_state.nos_data
    menu = sidebar_navigation()

    if menu == "Dashboard":
        dashboard_page(data)
    elif menu == "Calon Mahasiswa":
        students_page(data)
    elif menu == "Dokumen":
        documents_page(data)
    elif menu == "Invoice & Pembayaran":
        invoices_page(data)
    else:
        help_page()


if __name__ == "__main__":
    main()
